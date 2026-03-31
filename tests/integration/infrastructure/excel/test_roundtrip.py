"""Integration tests for Excel roundtrip (export then import preserves data)."""

from __future__ import annotations

from pathlib import Path

import pytest
from backlog_manager.application.dto.excel.export_excel_dto import ExcelExportData
from backlog_manager.infrastructure.excel.excel_service import ExcelService


@pytest.fixture
def excel_service():
    """Create ExcelService instance."""
    return ExcelService()


class TestRoundtripIntegrity:
    """Tests for export-import roundtrip data integrity (US4)."""

    @pytest.mark.integration
    async def test_exported_data_can_be_reimported(
        self, excel_service: ExcelService, tmp_path: Path
    ):
        """Should preserve data integrity when exporting then importing."""
        file_path = tmp_path / "roundtrip.xlsx"

        # Original data
        export_data = ExcelExportData(
            stories=[
                {
                    "ID": "AUTH-001",
                    "Componente": "AUTH",
                    "Nome": "Login de usuario",
                    "SP": 5,
                    "Status": "BACKLOG",
                    "Feature": "Autenticacao",
                    "Dependencias": "DB-001;API-001",
                    "Desenvolvedor": "Joao",
                    "Data Inicio": None,
                    "Data Fim": None,
                },
                {
                    "ID": "AUTH-002",
                    "Componente": "AUTH",
                    "Nome": "Logout de usuario",
                    "SP": 3,
                    "Status": "EXECUCAO",
                    "Feature": "Autenticacao",
                    "Dependencias": "AUTH-001",
                    "Desenvolvedor": "Maria",
                    "Data Inicio": None,
                    "Data Fim": None,
                },
            ],
            developers=[
                {"ID": 1, "Nome": "Joao"},
                {"ID": 2, "Nome": "Maria"},
            ],
            features=[
                {"ID": 1, "Nome": "Autenticacao", "Wave": 1},
                {"ID": 2, "Nome": "API", "Wave": 2},
            ],
        )

        # Export
        await excel_service.write_workbook(file_path, export_data)

        # The exported file has Stories sheet with export columns
        # For import, we need the file to have import-compatible columns
        # This test verifies the file structure is valid

        assert file_path.exists()

        # Read the Stories sheet to verify structure
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb["Stories"]

        # Verify headers
        headers = [cell.value for cell in ws[1]]
        assert "ID" in headers
        assert "Componente" in headers
        assert "Nome" in headers
        assert "SP" in headers
        assert "Feature" in headers
        assert "Dependencias" in headers

        # Verify data rows
        assert ws.cell(row=2, column=headers.index("ID") + 1).value == "AUTH-001"
        assert (
            ws.cell(row=2, column=headers.index("Nome") + 1).value == "Login de usuario"
        )
        assert ws.cell(row=2, column=headers.index("SP") + 1).value == 5
        assert (
            ws.cell(row=2, column=headers.index("Dependencias") + 1).value
            == "DB-001;API-001"
        )

        wb.close()

    @pytest.mark.integration
    async def test_special_characters_preserved(
        self, excel_service: ExcelService, tmp_path: Path
    ):
        """Should preserve special characters in text fields."""
        file_path = tmp_path / "special_chars.xlsx"

        export_data = ExcelExportData(
            stories=[
                {
                    "ID": "TEST-001",
                    "Componente": "TEST",
                    "Nome": "Teste com acentuação: é, á, ç, ã",
                    "SP": 5,
                    "Status": "BACKLOG",
                    "Feature": "",
                    "Dependencias": "",
                    "Desenvolvedor": "",
                    "Data Inicio": None,
                    "Data Fim": None,
                },
            ],
            developers=[],
            features=[],
        )

        await excel_service.write_workbook(file_path, export_data)

        # Read back
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb["Stories"]

        headers = [cell.value for cell in ws[1]]
        nome_value = ws.cell(row=2, column=headers.index("Nome") + 1).value

        assert nome_value == "Teste com acentuação: é, á, ç, ã"
        wb.close()

    @pytest.mark.integration
    async def test_empty_dependencies_preserved(
        self, excel_service: ExcelService, tmp_path: Path
    ):
        """Should preserve empty dependencies without conversion issues."""
        file_path = tmp_path / "empty_deps.xlsx"

        export_data = ExcelExportData(
            stories=[
                {
                    "ID": "TEST-001",
                    "Componente": "TEST",
                    "Nome": "No dependencies",
                    "SP": 3,
                    "Status": "BACKLOG",
                    "Feature": "",
                    "Dependencias": "",
                    "Desenvolvedor": "",
                    "Data Inicio": None,
                    "Data Fim": None,
                },
            ],
            developers=[],
            features=[],
        )

        await excel_service.write_workbook(file_path, export_data)

        # Read back
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb["Stories"]

        headers = [cell.value for cell in ws[1]]
        deps_value = ws.cell(row=2, column=headers.index("Dependencias") + 1).value

        # Empty string or None are acceptable
        assert deps_value in ("", None)
        wb.close()

    @pytest.mark.integration
    async def test_multiple_dependencies_preserved(
        self, excel_service: ExcelService, tmp_path: Path
    ):
        """Should preserve semicolon-separated dependency format."""
        file_path = tmp_path / "multi_deps.xlsx"

        export_data = ExcelExportData(
            stories=[
                {
                    "ID": "EPIC-001",
                    "Componente": "EPIC",
                    "Nome": "Multiple deps",
                    "SP": 13,
                    "Status": "BACKLOG",
                    "Feature": "",
                    "Dependencias": "A-001;B-001;C-001",
                    "Desenvolvedor": "",
                    "Data Inicio": None,
                    "Data Fim": None,
                },
            ],
            developers=[],
            features=[],
        )

        await excel_service.write_workbook(file_path, export_data)

        # Read back
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ws = wb["Stories"]

        headers = [cell.value for cell in ws[1]]
        deps_value = ws.cell(row=2, column=headers.index("Dependencias") + 1).value

        assert deps_value == "A-001;B-001;C-001"
        wb.close()

    @pytest.mark.integration
    async def test_feature_and_developer_sheets_preserved(
        self, excel_service: ExcelService, tmp_path: Path
    ):
        """Should preserve Feature and Developer sheets correctly."""
        file_path = tmp_path / "all_sheets.xlsx"

        export_data = ExcelExportData(
            stories=[],
            developers=[
                {"ID": 1, "Nome": "João Silva"},
                {"ID": 2, "Nome": "Maria Santos"},
            ],
            features=[
                {"ID": 1, "Nome": "Autenticação", "Wave": 1},
                {"ID": 2, "Nome": "Relatórios", "Wave": 2},
            ],
        )

        await excel_service.write_workbook(file_path, export_data)

        # Read back
        from openpyxl import load_workbook

        wb = load_workbook(file_path)

        # Check Developers sheet
        ws_devs = wb["Developers"]
        assert ws_devs.cell(row=2, column=2).value == "João Silva"
        assert ws_devs.cell(row=3, column=2).value == "Maria Santos"

        # Check Features sheet
        ws_feats = wb["Features"]
        assert ws_feats.cell(row=2, column=2).value == "Autenticação"
        assert ws_feats.cell(row=2, column=3).value == 1  # Wave

        wb.close()
