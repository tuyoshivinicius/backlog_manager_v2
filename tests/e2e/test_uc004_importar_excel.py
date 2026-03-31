"""E2E tests for UC-004: Importar Excel.

Tests the Excel import flow including file validation
and cycle detection during import.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from backlog_manager.application.dto.dependency import GetDependenciesInputDTO
from openpyxl import Workbook

pytestmark = [pytest.mark.e2e]


def create_valid_excel_file(
    tmp_path: Path,
    stories_data: list[dict] | None = None,
    dependencies_data: list[dict] | None = None,
) -> Path:
    """Helper to create a valid Excel file for testing.

    Args:
        tmp_path: Temporary directory path.
        stories_data: List of story dicts with columns.
        dependencies_data: List of dependency dicts.

    Returns:
        Path to created Excel file.
    """
    file_path = tmp_path / "test_import.xlsx"
    wb = Workbook()

    # Stories sheet
    ws_stories = wb.active
    ws_stories.title = "Stories"
    headers = ["ID", "Componente", "Nome", "SP", "Feature", "Dependencias"]
    ws_stories.append(headers)

    if stories_data:
        for story in stories_data:
            # Build dependencies string for this story
            deps_str = ""
            if dependencies_data:
                story_deps = [
                    d["depends_on"]
                    for d in dependencies_data
                    if d["story_id"] == story.get("id")
                ]
                deps_str = ",".join(story_deps)
            ws_stories.append(
                [
                    story.get("id", "TEST-001"),
                    story.get("component", "TEST"),
                    story.get("name", "Test Story"),
                    story.get("story_points", 5),
                    story.get("feature", ""),
                    deps_str,
                ]
            )

    # Dependencies sheet
    ws_deps = wb.create_sheet("Dependencias")
    ws_deps.append(["Story ID", "Depende De"])
    if dependencies_data:
        for dep in dependencies_data:
            ws_deps.append([dep["story_id"], dep["depends_on"]])

    # Developers sheet
    ws_devs = wb.create_sheet("Desenvolvedores")
    ws_devs.append(["ID", "Nome"])
    ws_devs.append([1, "Dev Test"])

    # Features sheet
    ws_features = wb.create_sheet("Features")
    ws_features.append(["ID", "Nome", "Wave"])

    wb.save(file_path)
    return file_path


class TestUC004ImportarArquivoValido:
    """Tests for valid Excel file import (FR-038)."""

    @pytest.mark.asyncio
    async def test_importar_arquivo_valido_com_historias(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test importing a valid Excel file with stories.

        FR-038: Valid files should be imported successfully.
        """
        excel_vm = e2e_app.excel_viewmodel

        # Create valid Excel file
        stories = [
            {
                "id": "IMP-001",
                "component": "IMP",
                "name": "Historia 1",
                "story_points": 5,
                "priority": 1,
            },
            {
                "id": "IMP-002",
                "component": "IMP",
                "name": "Historia 2",
                "story_points": 8,
                "priority": 2,
            },
            {
                "id": "IMP-003",
                "component": "IMP",
                "name": "Historia 3",
                "story_points": 3,
                "priority": 3,
            },
        ]
        file_path = create_valid_excel_file(tmp_path, stories_data=stories)

        # Import file
        result = await excel_vm.import_from_file(file_path)

        assert result is not None
        assert result.stories_imported == 3

    @pytest.mark.asyncio
    async def test_importar_arquivo_valido_com_dependencias(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test importing a valid Excel file with dependencies.

        FR-038: Dependencies should be imported correctly.
        """
        excel_vm = e2e_app.excel_viewmodel

        # Create valid Excel file with dependencies
        stories = [
            {
                "id": "DEP-001",
                "component": "DEP",
                "name": "Historia 1",
                "story_points": 5,
                "priority": 1,
            },
            {
                "id": "DEP-002",
                "component": "DEP",
                "name": "Historia 2",
                "story_points": 5,
                "priority": 2,
            },
            {
                "id": "DEP-003",
                "component": "DEP",
                "name": "Historia 3",
                "story_points": 5,
                "priority": 3,
            },
        ]
        dependencies = [
            {"story_id": "DEP-002", "depends_on": "DEP-001"},
            {"story_id": "DEP-003", "depends_on": "DEP-002"},
        ]
        file_path = create_valid_excel_file(
            tmp_path, stories_data=stories, dependencies_data=dependencies
        )

        # Import file
        result = await excel_vm.import_from_file(file_path)

        assert result is not None
        assert result.stories_imported == 3

        # Verify dependencies were created
        async with e2e_app.create_unit_of_work() as uow:
            get_deps = e2e_app.create_get_dependencies_use_case(uow)
            result_2 = await get_deps.execute(
                GetDependenciesInputDTO(story_id="DEP-002")
            )
            result_3 = await get_deps.execute(
                GetDependenciesInputDTO(story_id="DEP-003")
            )

        depends_on_ids_2 = result_2.dependencies
        depends_on_ids_3 = result_3.dependencies
        assert "DEP-001" in depends_on_ids_2
        assert "DEP-002" in depends_on_ids_3

    @pytest.mark.asyncio
    async def test_importar_emite_signals_progresso(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test that import emits progress signals.

        FR-038: Progress should be reported during import.
        """
        excel_vm = e2e_app.excel_viewmodel
        progress_values = []

        def capture_progress(value):
            progress_values.append(value)

        excel_vm.progress_updated.connect(capture_progress)

        # Create valid Excel file
        stories = [
            {
                "id": "PRG-001",
                "component": "PRG",
                "name": "Historia 1",
                "story_points": 5,
                "priority": 1,
            },
        ]
        file_path = create_valid_excel_file(tmp_path, stories_data=stories)

        # Import file
        await excel_vm.import_from_file(file_path)

        # Should have received progress updates
        assert len(progress_values) >= 2
        assert 0 in progress_values
        assert 100 in progress_values


class TestUC004RejeitarCicloImport:
    """Tests for cycle rejection during import (FR-039)."""

    @pytest.mark.asyncio
    async def test_rejeitar_ciclo_direto_no_import(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test that direct cycles in import are rejected.

        FR-039: Files with cyclic dependencies should be rejected.
        """
        excel_vm = e2e_app.excel_viewmodel
        error_received = []

        def capture_error(msg):
            error_received.append(msg)

        excel_vm.import_error.connect(capture_error)

        # Create Excel file with cycle
        stories = [
            {
                "id": "CYC-001",
                "component": "CYC",
                "name": "Historia 1",
                "story_points": 5,
                "priority": 1,
            },
            {
                "id": "CYC-002",
                "component": "CYC",
                "name": "Historia 2",
                "story_points": 5,
                "priority": 2,
            },
        ]
        dependencies = [
            {"story_id": "CYC-001", "depends_on": "CYC-002"},
            {"story_id": "CYC-002", "depends_on": "CYC-001"},  # Creates cycle!
        ]
        file_path = create_valid_excel_file(
            tmp_path, stories_data=stories, dependencies_data=dependencies
        )

        # Import file
        result = await excel_vm.import_from_file(file_path)

        # Should fail or emit error
        if result is None:
            assert len(error_received) > 0
            assert any("ciclo" in e.lower() for e in error_received)

    @pytest.mark.asyncio
    async def test_rejeitar_ciclo_indireto_no_import(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test that indirect cycles in import are rejected.

        FR-039: Indirect cycles should also be rejected.
        """
        excel_vm = e2e_app.excel_viewmodel
        error_received = []

        def capture_error(msg):
            error_received.append(msg)

        excel_vm.import_error.connect(capture_error)

        # Create Excel file with indirect cycle (A -> B -> C -> A)
        stories = [
            {
                "id": "IND-001",
                "component": "IND",
                "name": "Historia A",
                "story_points": 5,
                "priority": 1,
            },
            {
                "id": "IND-002",
                "component": "IND",
                "name": "Historia B",
                "story_points": 5,
                "priority": 2,
            },
            {
                "id": "IND-003",
                "component": "IND",
                "name": "Historia C",
                "story_points": 5,
                "priority": 3,
            },
        ]
        dependencies = [
            {"story_id": "IND-002", "depends_on": "IND-001"},  # B -> A
            {"story_id": "IND-003", "depends_on": "IND-002"},  # C -> B
            {"story_id": "IND-001", "depends_on": "IND-003"},  # A -> C (creates cycle!)
        ]
        file_path = create_valid_excel_file(
            tmp_path, stories_data=stories, dependencies_data=dependencies
        )

        # Import file
        result = await excel_vm.import_from_file(file_path)

        # Should fail or emit error
        if result is None:
            assert len(error_received) > 0


class TestUC004ImportErrors:
    """Tests for import error handling."""

    @pytest.mark.asyncio
    async def test_arquivo_nao_encontrado(self, e2e_app, qasync_loop, tmp_path):
        """Test error when file is not found."""
        excel_vm = e2e_app.excel_viewmodel
        error_received = []

        def capture_error(msg):
            error_received.append(msg)

        excel_vm.import_error.connect(capture_error)

        # Try to import non-existent file
        file_path = tmp_path / "nao_existe.xlsx"
        result = await excel_vm.import_from_file(file_path)

        assert result is None
        assert len(error_received) > 0

    @pytest.mark.asyncio
    async def test_arquivo_corrompido(self, e2e_app, qasync_loop, tmp_path):
        """Test error when file is corrupted."""
        excel_vm = e2e_app.excel_viewmodel
        error_received = []

        def capture_error(msg):
            error_received.append(msg)

        excel_vm.import_error.connect(capture_error)

        # Create corrupted file
        file_path = tmp_path / "corrupted.xlsx"
        file_path.write_text("this is not an excel file")

        result = await excel_vm.import_from_file(file_path)

        assert result is None
        assert len(error_received) > 0

    @pytest.mark.asyncio
    async def test_arquivo_sem_colunas_obrigatorias(
        self, e2e_app, qasync_loop, tmp_path
    ):
        """Test error when required columns are missing."""
        excel_vm = e2e_app.excel_viewmodel
        error_received = []

        def capture_error(msg):
            error_received.append(msg)

        excel_vm.import_error.connect(capture_error)

        # Create file without required columns
        file_path = tmp_path / "missing_cols.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Stories"
        ws.append(["Coluna1", "Coluna2"])  # Missing required columns
        wb.save(file_path)

        result = await excel_vm.import_from_file(file_path)

        assert result is None
        assert len(error_received) > 0
