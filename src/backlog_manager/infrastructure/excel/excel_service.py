"""Excel I/O service implementation using openpyxl."""

import asyncio
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backlog_manager.application.dto.excel.export_excel_dto import ExcelExportData
from backlog_manager.application.dto.excel.import_excel_dto import ExcelReadResult
from backlog_manager.application.exceptions.excel_exceptions import (
    ExcelFileCorruptedException,
    ExcelFileNotFoundException,
    ExcelMissingHeaderException,
    ExcelPermissionException,
)

logger = logging.getLogger(__name__)

REQUIRED_HEADERS = ["ID", "Componente", "Nome", "SP", "Feature", "Dependencias"]


class ExcelService:
    """Servico de infraestrutura para operacoes Excel via openpyxl.

    Implementa ExcelServiceProtocol definido na camada Application.
    Utiliza asyncio.to_thread() para nao bloquear o event loop
    durante operacoes de I/O com arquivos Excel.
    """

    async def read_stories_from_file(self, file_path: Path) -> ExcelReadResult:
        """Le arquivo Excel de forma assincrona.

        Args:
            file_path: Caminho absoluto para arquivo .xlsx.

        Returns:
            ExcelReadResult com lista de dicionarios (linhas) e warnings.

        Raises:
            ExcelFileNotFoundException: Arquivo nao encontrado.
            ExcelFileCorruptedException: Arquivo invalido ou corrompido.
            ExcelMissingHeaderException: Coluna obrigatoria ausente.
            ExcelPermissionException: Sem permissao de leitura.
        """
        return await asyncio.to_thread(self._read_stories_sync, file_path)

    def _read_stories_sync(self, file_path: Path) -> ExcelReadResult:
        """Operacao sincrona de leitura - executada em thread separada.

        Args:
            file_path: Caminho absoluto para arquivo .xlsx.

        Returns:
            ExcelReadResult com dados lidos do arquivo.

        Raises:
            ExcelFileNotFoundException: Arquivo nao encontrado.
            ExcelFileCorruptedException: Arquivo invalido ou corrompido.
            ExcelMissingHeaderException: Coluna obrigatoria ausente.
            ExcelPermissionException: Sem permissao de leitura.
        """
        logger.info("Lendo arquivo Excel: %s", file_path)
        self._validate_file_path(file_path)
        wb = self._open_workbook(file_path)

        try:
            ws = self._get_active_worksheet(wb, file_path)
            headers = self._read_and_validate_headers(ws, file_path)
            rows = self._read_data_rows(ws, headers)

            logger.info(
                "Leitura concluida: %d linhas lidas de %s", len(rows), file_path
            )
            return ExcelReadResult(rows=rows, warnings=[])

        finally:
            wb.close()

    @staticmethod
    def _validate_file_path(file_path: Path) -> None:
        """Valida existencia e extensao do arquivo.

        Args:
            file_path: Caminho do arquivo.

        Raises:
            ExcelFileNotFoundException: Arquivo nao encontrado.
            ExcelFileCorruptedException: Extensao invalida.
        """
        if not file_path.exists():
            logger.error("Arquivo nao encontrado: %s", file_path)
            raise ExcelFileNotFoundException(f"Arquivo nao encontrado: {file_path}")

        if file_path.suffix.lower() != ".xlsx":
            logger.error("Formato nao suportado: %s", file_path.suffix)
            raise ExcelFileCorruptedException(
                "Formato de arquivo nao suportado. Use .xlsx"
            )

    @staticmethod
    def _open_workbook(file_path: Path) -> Workbook:
        """Abre workbook com tratamento de erros.

        Args:
            file_path: Caminho do arquivo.

        Returns:
            Workbook aberto.

        Raises:
            ExcelPermissionException: Sem permissao de leitura.
            ExcelFileCorruptedException: Arquivo invalido ou corrompido.
        """
        try:
            return load_workbook(file_path, read_only=True, data_only=True)
        except PermissionError as e:
            logger.error("Sem permissao para ler arquivo: %s", file_path)
            raise ExcelPermissionException(
                f"Sem permissao para ler arquivo: {file_path}"
            ) from e
        except InvalidFileException as e:
            logger.error("Arquivo corrompido ou invalido: %s", file_path)
            raise ExcelFileCorruptedException(
                "Arquivo invalido ou corrompido. Verifique o formato"
            ) from e
        except Exception as e:
            logger.error("Erro ao abrir arquivo Excel: %s - %s", file_path, str(e))
            raise ExcelFileCorruptedException(
                "Arquivo invalido ou corrompido. Verifique o formato"
            ) from e

    @staticmethod
    def _get_active_worksheet(wb: Workbook, file_path: Path) -> Any:
        """Obtem a worksheet ativa do workbook.

        Args:
            wb: Workbook aberto.
            file_path: Caminho do arquivo (para mensagens de erro).

        Returns:
            Worksheet ativa.

        Raises:
            ExcelFileCorruptedException: Nenhuma planilha ativa.
        """
        ws = wb.active
        if ws is None:
            logger.error("Arquivo nao contem planilha ativa: %s", file_path)
            raise ExcelFileCorruptedException("Arquivo Excel nao contem planilha ativa")
        return ws

    @staticmethod
    def _read_and_validate_headers(ws: Any, file_path: Path) -> list[str]:
        """Le e valida headers da primeira linha.

        Args:
            ws: Worksheet ativa.
            file_path: Caminho do arquivo (para mensagens de erro).

        Returns:
            Lista de nomes dos headers.

        Raises:
            ExcelMissingHeaderException: Header ausente ou arquivo vazio.
        """
        first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not first_row or not first_row[0]:
            logger.error("Arquivo vazio ou sem headers: %s", file_path)
            raise ExcelMissingHeaderException(
                "Arquivo Excel vazio ou sem linha de cabecalho"
            )

        headers = [str(cell) if cell is not None else "" for cell in first_row[0]]

        for required in REQUIRED_HEADERS:
            if required not in headers:
                logger.error(
                    "Coluna obrigatoria ausente: %s no arquivo %s",
                    required,
                    file_path,
                )
                raise ExcelMissingHeaderException(
                    f"Coluna obrigatoria '{required}' nao encontrada na linha 1"
                )

        return headers

    @staticmethod
    def _read_data_rows(ws: Any, headers: list[str]) -> list[dict[str, Any]]:
        """Le linhas de dados da worksheet.

        Args:
            ws: Worksheet ativa.
            headers: Lista de nomes dos headers.

        Returns:
            Lista de dicionarios representando cada linha.
        """
        rows: list[dict[str, Any]] = []

        for _row_idx, row_values in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not any(cell is not None and str(cell).strip() for cell in row_values):
                continue

            row_dict: dict[str, Any] = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row_values):
                    row_dict[header] = row_values[col_idx]
                else:
                    row_dict[header] = None

            rows.append(row_dict)

        return rows

    async def write_workbook(self, file_path: Path, data: ExcelExportData) -> None:
        """Escreve arquivo Excel de forma assincrona.

        Args:
            file_path: Caminho absoluto para arquivo .xlsx de destino.
            data: Dados a serem escritos (stories, developers, features).

        Raises:
            ExcelPermissionException: Sem permissao de escrita.
            ExcelFileCorruptedException: Erro ao escrever arquivo.
        """
        await asyncio.to_thread(self._write_workbook_sync, file_path, data)

    def _write_workbook_sync(self, file_path: Path, data: ExcelExportData) -> None:
        """Operacao sincrona de escrita - executada em thread separada.

        Args:
            file_path: Caminho absoluto para arquivo .xlsx de destino.
            data: Dados a serem escritos.

        Raises:
            ExcelPermissionException: Sem permissao de escrita.
            ExcelFileCorruptedException: Erro ao escrever arquivo.
        """
        logger.info("Escrevendo arquivo Excel: %s", file_path)

        try:
            wb = Workbook()
            # Remove default sheet created by openpyxl
            if wb.active is not None:
                wb.remove(wb.active)

            # Stories sheet
            self._write_sheet(wb, "Stories", data.stories)

            # Developers sheet
            self._write_sheet(wb, "Developers", data.developers)

            # Features sheet
            self._write_sheet(wb, "Features", data.features)

            wb.save(file_path)
            logger.info(
                "Export concluido: %d historias, %d desenvolvedores, %d features em %s",
                len(data.stories),
                len(data.developers),
                len(data.features),
                file_path,
            )

        except PermissionError as e:
            logger.error("Sem permissao para escrever arquivo: %s", file_path)
            raise ExcelPermissionException(
                f"Sem permissao para escrever arquivo: {file_path}"
            ) from e
        except Exception as e:
            logger.error("Erro ao escrever arquivo Excel: %s - %s", file_path, str(e))
            raise ExcelFileCorruptedException(
                f"Erro ao criar arquivo Excel: {e}"
            ) from e

    def _write_sheet(self, wb: Workbook, name: str, data: list[dict[str, Any]]) -> None:
        """Cria worksheet com dados.

        Args:
            wb: Workbook destino.
            name: Nome da worksheet.
            data: Lista de dicionarios com dados das linhas.
        """
        ws = wb.create_sheet(name)

        if data:
            # Write headers from first row's keys
            headers = list(data[0].keys())
            ws.append(headers)

            # Write data rows
            for row_dict in data:
                row_values = [row_dict.get(h) for h in headers]
                ws.append(row_values)
        else:
            # Empty sheet - just write headers based on sheet type
            if name == "Stories":
                ws.append(
                    [
                        "ID",
                        "Componente",
                        "Nome",
                        "SP",
                        "Status",
                        "Feature",
                        "Dependencias",
                        "Desenvolvedor",
                        "Data Inicio",
                        "Data Fim",
                    ]
                )
            elif name == "Developers":
                ws.append(["ID", "Nome"])
            elif name == "Features":
                ws.append(["ID", "Nome", "Wave"])
